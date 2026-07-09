#!/usr/bin/env python3
"""skill_gap_xlsx.py — regenerate the human copy of the skill-gap from its JSON.

The skill-gap skill keeps two artifacts in the user's project folder:

  * career-graph/skill-gap.json   the machine copy — levels, ranked gaps, plan,
                                  progress over time. Source of truth.
  * career-graph/skill-gap.xlsx   the human copy — a styled spreadsheet the user
                                  can open, skim, and even edit by hand.

This script reads the JSON and (re)writes the spreadsheet so the two never drift.
It is meant to run on every skill-gap session: JSON is authored by the agent,
xlsx is always derived from it.

Columns, in this exact order:
  Skill · Your level · Target level · Gap · Priority · Plan · How to prove it ·
  Time est. · Status · Last updated

Graceful degradation: openpyxl is the repo's first non-stdlib dependency. If it
isn't installed we fall back to writing career-graph/skill-gap.csv with the same
columns via the stdlib csv module, and print a one-line note on how to get the
real xlsx (`pip install openpyxl`). The skill never hard-fails for lack of a
spreadsheet library.

Usage:
  scripts/skill_gap_xlsx.py [project-folder]

  project-folder   the folder the user runs Hope in (default: current dir).
                   Reads <project-folder>/career-graph/skill-gap.json and writes
                   the .xlsx (or .csv fallback) alongside it.

Exit code: 0 = spreadsheet (or csv fallback) written, 2 = JSON missing/unreadable.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

# Project-folder relative — the folder the user runs Hope in, NOT a home dir.
# Mirrors graph_query.DEFAULT_GRAPH_PATH so both scripts read the same career-graph/.
GRAPH_DIR = "career-graph"
JSON_NAME = "skill-gap.json"
XLSX_NAME = "skill-gap.xlsx"
CSV_NAME = "skill-gap.csv"

# Column order is a locked contract — the skill's report and the JSON authoring
# both assume these headers in this sequence. Do not reorder without sign-off.
COLUMNS: list[str] = [
    "Skill",
    "Your level",
    "Target level",
    "Gap",
    "Priority",
    "Plan",
    "How to prove it",
    "Time est.",
    "Status",
    "Last updated",
]

# Sensible widths (in openpyxl column-width units ≈ characters) per column.
# Prose columns get room; level/priority/status stay compact.
COLUMN_WIDTHS: dict[str, int] = {
    "Skill": 22,
    "Your level": 14,
    "Target level": 14,
    "Gap": 10,
    "Priority": 12,
    "Plan": 48,
    "How to prove it": 48,
    "Time est.": 12,
    "Status": 14,
    "Last updated": 16,
}

# Map each JSON skill object's keys to the spreadsheet columns. The JSON authored
# by the skill uses these snake_case keys; read defensively so a partial/edited
# JSON still produces a row rather than crashing.
FIELD_FOR_COLUMN: dict[str, str] = {
    "Skill": "skill",
    "Your level": "your_level",
    "Target level": "target_level",
    "Gap": "gap",
    "Priority": "priority",
    "Plan": "plan",
    "How to prove it": "how_to_prove_it",
    "Time est.": "time_est",
    "Status": "status",
    "Last updated": "last_updated",
}


def load_skill_gap(json_path: Path) -> list[dict[str, Any]]:
    """Read skill-gap.json and return the list of skill rows, defensively.

    Accepts either a top-level list of skill objects, or an object with a
    "skills" (or "rows") list. Anything else yields an empty list rather than
    raising — a hand-edited JSON should never hard-crash the regeneration.
    """
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = data.get("skills") or data.get("rows") or []
    else:
        rows = []

    return [r for r in rows if isinstance(r, dict)]


def cell_value(row: dict[str, Any], column: str) -> str:
    """Pull the value for a column from a skill row, coercing to a clean string.

    LLM/hand-authored JSON may hand us None, numbers, or lists — coerce every
    field to a display string so neither openpyxl nor csv chokes on it.
    """
    raw = row.get(FIELD_FOR_COLUMN[column])
    if raw is None:
        return ""
    if isinstance(raw, (list, tuple)):
        return ", ".join(str(item) for item in raw)
    if isinstance(raw, bool):
        return "yes" if raw else "no"
    return str(raw)


def write_xlsx(rows: list[dict[str, Any]], xlsx_path: Path) -> None:
    """Write the styled spreadsheet via openpyxl. Caller handles ImportError."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Skill gap"

    # Header row — bold, frozen so it stays visible while scrolling.
    ws.append(COLUMNS)
    header_font = Font(bold=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).font = header_font
    ws.freeze_panes = "A2"

    # Data rows.
    for row in rows:
        ws.append([cell_value(row, column) for column in COLUMNS])

    # Column widths + wrap on the prose columns so plans/proofs stay readable.
    wrap = Alignment(vertical="top", wrap_text=True)
    for idx, column in enumerate(COLUMNS, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS[column]
        if column in ("Plan", "How to prove it"):
            for cell in ws[letter]:
                cell.alignment = wrap

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def write_csv(rows: list[dict[str, Any]], csv_path: Path) -> None:
    """Fallback writer — same columns, stdlib csv, no styling."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        for row in rows:
            writer.writerow([cell_value(row, column) for column in COLUMNS])


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Regenerate career-graph/skill-gap.xlsx from skill-gap.json "
                    "(falls back to skill-gap.csv if openpyxl is unavailable)."
    )
    parser.add_argument(
        "project_folder",
        type=Path,
        nargs="?",
        default=Path.cwd(),
        help="folder the user runs Hope in (default: current directory)",
    )
    args = parser.parse_args(argv)

    graph_dir = args.project_folder / GRAPH_DIR
    json_path = graph_dir / JSON_NAME
    xlsx_path = graph_dir / XLSX_NAME
    csv_path = graph_dir / CSV_NAME

    if not json_path.exists():
        print(f"ERROR: {json_path} not found — run the skill-gap skill first to "
              f"create it.", file=sys.stderr)
        return 2

    try:
        rows = load_skill_gap(json_path)
    except (json.JSONDecodeError, OSError) as exc:
        print(f"ERROR: could not read {json_path}: {exc}", file=sys.stderr)
        return 2

    try:
        write_xlsx(rows, xlsx_path)
    except ImportError:
        write_csv(rows, csv_path)
        print(f"Wrote {csv_path} ({len(rows)} skills).")
        print("Note: openpyxl isn't installed, so I wrote a .csv instead of a "
              ".xlsx. For the styled spreadsheet, run 'pip install openpyxl' and "
              "regenerate.")
        return 0

    print(f"Wrote {xlsx_path} ({len(rows)} skills).")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
